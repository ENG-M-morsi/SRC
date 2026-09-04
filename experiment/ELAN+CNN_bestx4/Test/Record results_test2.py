import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Parse log.txt ---
log_file = os.path.join(os.path.dirname(__file__), "log.txt")

with open(log_file, "r") as f:
    content = f.read()

# Extract total parameters
params_match = re.search(r"([\d,]+) total parameters", content)
total_params = params_match.group(1).replace(",", "") if params_match else "?"
params_k = f"{int(total_params) // 1000:,}K" if total_params != "?" else "?"

# Extract PSNR and SSIM for each dataset (supports any scale like x2, x3, x4, x8, etc.)
datasets = ["Set5", "Set14", "B100", "Urban100", "Manga109"]
results = {}
scale_detected = None

for ds in datasets:
    # Pattern now matches x2, x3, x4, x8, or any number
    pattern = rf"\[{ds} x(\d+)\]\s+Average PSNR:\s*([\d.]+) dB \| SSIM:\s*([\d.]+)"
    m = re.search(pattern, content)
    if m:
        scale = int(m.group(1))
        psnr = float(m.group(2))
        ssim = float(m.group(3))
        results[ds] = {"psnr": psnr, "ssim": ssim}
        if scale_detected is None:
            scale_detected = scale

# Extract model name (first word before '(')
model_match = re.match(r"^(\w+)\(", content.strip())
model_name = model_match.group(1) if model_match else "Model"

# Add scale to model name if detected
if scale_detected:
    model_name = f"{model_name} x{scale_detected}"

# --- Build Excel ---
wb = Workbook()
ws = wb.active
ws.title = "Results"

# Colors
RED     = "FF0000"
BLACK   = "000000"
WHITE   = "FFFFFF"
DARK_BG = "1F1F1F"
GRAY_BG = "D9D9D9"
HEADER_BG = "2F2F2F"

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(cell, bold=False, font_color=BLACK, bg_color=None,
               h_align="center", v_align="center", font_size=10):
    cell.font = Font(name="Arial", bold=bold, color=font_color, size=font_size)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=True)
    if bg_color:
        cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.border = border

# ---- Headers ----
headers = ["Model", "Set5", "Set14", "B100", "Urban100", "Manga109", "Average"]
col_widths = [14, 14, 14, 14, 14, 14, 14]

# Row 1: Main headers
for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell_style(cell, bold=True, font_color=WHITE, bg_color=HEADER_BG)
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# Row 2: Sub-headers
ws.cell(row=2, column=1, value="")
ws["A2"].fill = PatternFill("solid", fgColor=HEADER_BG)
ws["A2"].border = border
for col_idx in range(2, 8):
    cell = ws.cell(row=2, column=col_idx, value="PSNR(dB)/\nSSIM")
    cell_style(cell, bold=True, font_color=RED, bg_color=HEADER_BG)

# ---- Compute averages ----
psnr_vals = [results[ds]["psnr"] for ds in datasets if ds in results]
ssim_vals = [results[ds]["ssim"] for ds in datasets if ds in results]
if psnr_vals and ssim_vals:
    avg_psnr = sum(psnr_vals) / len(psnr_vals)
    avg_ssim = sum(ssim_vals) / len(ssim_vals)
    avg_text = f"{avg_psnr:.3f}/\n{avg_ssim:.4f}"
else:
    avg_text = "N/A"

# ---- Row 3: Data ----
ws.cell(row=3, column=1, value=f"\n{params_k}")
cell_style(ws.cell(row=3, column=1), bold=True, font_color=BLACK, bg_color=GRAY_BG)

for col_idx, ds in enumerate(datasets, start=2):
    if ds in results:
        psnr = results[ds]["psnr"]
        ssim = results[ds]["ssim"]
        val = f"{psnr:.3f}/\n{ssim:.4f}"
    else:
        val = "N/A"
    cell = ws.cell(row=3, column=col_idx, value=val)
    cell_style(cell, bold=False, font_color=RED, bg_color=GRAY_BG)

# ---- Average column ----
cell_avg = ws.cell(row=3, column=7, value=avg_text)
cell_style(cell_avg, bold=False, font_color=RED, bg_color=GRAY_BG)

# Row heights
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 30
ws.row_dimensions[3].height = 35

output_path = os.path.join(os.path.dirname(__file__), "results_table.xlsx")
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Model: {model_name} | Params: {params_k} | Scale: x{scale_detected if scale_detected else '?'}")
for ds, v in results.items():
    print(f"  {ds}: PSNR={v['psnr']} | SSIM={v['ssim']}")
if psnr_vals and ssim_vals:
    print(f"  Average: PSNR={avg_psnr:.3f} | SSIM={avg_ssim:.4f}")