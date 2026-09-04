import os
import re
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time

# === Helper: get folder name and model name ===
base_dir = os.path.dirname(os.path.abspath(__file__))
folder_name = os.path.basename(base_dir)

# Try to infer model name from log.txt (first line)
model_name = "unknown"
log_file_path = os.path.join(base_dir, "log.txt")
if os.path.exists(log_file_path):
    with open(log_file_path, "r", encoding="utf-8") as f:
        first_line = f.readline().strip()
        model_match = re.search(r'(HUTCN|DHTCUN|RFDN|EDSR|RCAN|SwinT|P_HTCB)', first_line, re.IGNORECASE)
        if model_match:
            model_name = model_match.group(1)
        else:
            model_name = "DHTCUN"

output_excel_base = os.path.join(base_dir, f"results_{model_name}_{folder_name}.xlsx")

# === Regex patterns (FIXED: allow digits in loss name) ===
epoch_lr_pattern = re.compile(r"\[Epoch\s+(\d+)\].*Learning\s+rate:\s*([\d.e+\-]+)", re.IGNORECASE)
loss_pair_pattern = re.compile(r"\[([A-Za-z0-9_]+):\s*([\d.e+\-]+)\]")   # now matches L1, L2, Perceptual_Loss, etc.
psnr_pattern = re.compile(r"Average\s+PSNR:\s*([\d.]+)")
best_psnr_pattern = re.compile(r"\(Best:\s*([\d.]+)\s*@epoch\s+(\d+)\)")

# === Data structures ===
epoch_losses = defaultdict(lambda: defaultdict(list))
epoch_lr = {}
epoch_psnr = {}
epoch_is_best = set()
best_psnr_overall = -np.inf
best_epoch_overall = None
current_epoch = None

# === Parse log ===
print("Parsing log file...")
with open(log_file_path, "r", encoding="utf-8") as f:
    for line_num, line in enumerate(f, 1):
        line = line.strip()
        if not line:
            continue
        
        # 1. Find epoch and learning rate
        m_epoch_lr = epoch_lr_pattern.search(line)
        if m_epoch_lr:
            current_epoch = int(m_epoch_lr.group(1))
            epoch_lr[current_epoch] = float(m_epoch_lr.group(2))
            # print(f"Found epoch {current_epoch} with LR {epoch_lr[current_epoch]}")
        
        # 2. Find loss pairs like [L1: 0.1339] (now includes numbers)
        loss_pairs = loss_pair_pattern.findall(line)
        if loss_pairs and current_epoch is not None:
            for loss_name, loss_val in loss_pairs:
                try:
                    epoch_losses[current_epoch][loss_name].append(float(loss_val))
                except ValueError:
                    pass
        
        # 3. Find PSNR
        m_psnr = psnr_pattern.search(line)
        if m_psnr and current_epoch is not None:
            epoch_psnr[current_epoch] = float(m_psnr.group(1))
        
        # 4. Find best PSNR indicators
        m_best = best_psnr_pattern.search(line)
        if m_best:
            best_val = float(m_best.group(1))
            best_ep = int(m_best.group(2))
            epoch_is_best.add(best_ep)
            if best_val > best_psnr_overall:
                best_psnr_overall = best_val
                best_epoch_overall = best_ep

print(f"Found {len(epoch_lr)} epochs, {len(epoch_psnr)} PSNR records, {len(epoch_losses)} loss records.")

# Build DataFrame
all_epochs = sorted(set(epoch_losses.keys()) | set(epoch_lr.keys()) | set(epoch_psnr.keys()))
data_rows = []
for ep in all_epochs:
    row = {'Epoch': ep}
    row['Learning_Rate'] = epoch_lr.get(ep, np.nan)
    row['PSNR'] = epoch_psnr.get(ep, np.nan)
    loss_dict = epoch_losses.get(ep, {})
    for loss_name, vals in loss_dict.items():
        row[loss_name] = np.mean(vals) if vals else np.nan
    data_rows.append(row)

df = pd.DataFrame(data_rows)

# Identify loss columns (all columns except Epoch, Learning_Rate, PSNR)
loss_columns = [col for col in df.columns if col not in ['Epoch', 'Learning_Rate', 'PSNR']]
loss_columns.sort()

# Add Total_Loss column (sum of all individual losses per epoch)
if len(loss_columns) > 0:
    # If only one loss, Total_Loss = that loss (for consistency)
    if len(loss_columns) == 1:
        df['Total_Loss'] = df[loss_columns[0]]
    else:
        df['Total_Loss'] = df[loss_columns].sum(axis=1, skipna=False)
    # Reorder columns: Epoch, Learning_Rate, individual losses, Total_Loss, PSNR
    ordered_columns = ['Epoch', 'Learning_Rate'] + loss_columns + ['Total_Loss', 'PSNR']
else:
    ordered_columns = ['Epoch', 'Learning_Rate', 'PSNR']

df = df[ordered_columns]

print(f"Loss columns found: {loss_columns}")
print(f"Total_Loss column added (sum of losses)")

# === Save to Excel with coloring ===
output_excel = output_excel_base
attempt = 1
max_attempts = 5
while attempt <= max_attempts:
    try:
        df.to_excel(output_excel, index=False, sheet_name='Training Results')
        break
    except PermissionError:
        if attempt == 1:
            print(f"⚠️ Permission denied: {output_excel} is open. Please close the file.")
        name, ext = os.path.splitext(output_excel_base)
        output_excel = f"{name}_{attempt}{ext}"
        print(f"Trying alternative filename: {output_excel}")
        attempt += 1
        time.sleep(0.5)
    except Exception as e:
        print(f"Unexpected error: {e}")
        exit()

if attempt <= max_attempts:
    try:
        wb = load_workbook(output_excel)
        ws = wb['Training Results']
        # Find PSNR column index
        psnr_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'PSNR':
                psnr_col_idx = col_idx
                break
        intermediate_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        if psnr_col_idx:
            for ep in epoch_is_best:
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == ep:
                        ws.cell(row=row, column=psnr_col_idx).fill = intermediate_fill
                        break
            if best_epoch_overall is not None:
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == best_epoch_overall:
                        ws.cell(row=row, column=psnr_col_idx).fill = best_fill
                        break
        wb.save(output_excel)
        print(f"✅ Results saved to: {output_excel} with best PSNR highlighted")
    except Exception as e:
        print(f"Warning: Could not apply styling: {e}")

# === Plotting with grid (all plots saved as PDF) ===
grid_style = {'linestyle': '--', 'alpha': 0.6, 'color': 'gray'}

# 1. PSNR plot
if not df['PSNR'].isna().all():
    plt.figure(figsize=(10, 6))
    plt.plot(df['Epoch'], df['PSNR'], linestyle='-', color='b', label='PSNR')
    if best_epoch_overall is not None:
        plt.scatter(best_epoch_overall, best_psnr_overall, color='red', s=100, zorder=5,
                    label=f'Best PSNR = {best_psnr_overall:.3f} dB')
        plt.annotate(f'{best_psnr_overall:.3f}', (best_epoch_overall, best_psnr_overall),
                     textcoords="offset points", xytext=(10,10), ha='center')
    plt.xlabel("Epoch")
    plt.ylabel("PSNR (dB)")
    plt.title("PSNR vs Epochs")
    plt.grid(True, **grid_style)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(base_dir, "PSNR_plot.pdf"))
    print("📈 PSNR curve saved as PDF")

# 2. Individual loss plots (including Total_Loss)
all_loss_columns = loss_columns + (['Total_Loss'] if 'Total_Loss' in df.columns else [])
for loss_name in all_loss_columns:
    loss_vals = df[loss_name]
    if not loss_vals.isna().all():
        plt.figure(figsize=(10, 6))
        plt.plot(df['Epoch'], loss_vals, linestyle='-', label=loss_name)
        plt.xlabel("Epoch")
        plt.ylabel(f"{loss_name} Loss")
        plt.title(f"{loss_name} vs Epochs")
        plt.yscale('log')   # log scale for better visualization
        plt.grid(True, **grid_style)
        plt.legend()
        plt.tight_layout()
        plt.savefig(os.path.join(base_dir, f"{loss_name}_loss_plot.pdf"))
        print(f"📉 {loss_name} loss curve saved as PDF")

# 3. Combined loss plot (all losses + Total_Loss)
if all_loss_columns:
    plt.figure(figsize=(12, 7))
    line_styles = ['-', '--', '-.', ':']
    colors = plt.cm.tab10(np.linspace(0, 1, len(all_loss_columns)))
    for idx, loss_name in enumerate(all_loss_columns):
        loss_vals = df[loss_name]
        if not loss_vals.isna().all():
            style = line_styles[idx % len(line_styles)]
            plt.plot(df['Epoch'], loss_vals, linestyle=style, color=colors[idx], label=loss_name)
    plt.xlabel("Epoch")
    plt.ylabel("Loss (log scale)")
    plt.title("All Losses (including Total) vs Epochs")
    plt.yscale('log')
    plt.grid(True, **grid_style)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(base_dir, "All_losses_combined.pdf"))
    print("📊 Combined loss curves saved as PDF")

print("\n✅ Processing complete.")
print(f"Best overall PSNR: {best_psnr_overall:.3f} dB at epoch {best_epoch_overall}")
print(f"Number of intermediate best epochs: {len(epoch_is_best)}")